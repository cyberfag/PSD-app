# coding=utf-8

from tkinter import filedialog
from tkinter import *
from collections import OrderedDict as od

import numpy as np
import openpyxl
import pandas as pd
import xlsxwriter
from neo import PlexonIO
from scipy.signal import welch
from tkinter import messagebox as mb


class app():
    def __init__(self):
        self.smell = 1
        self.start = 10
        self.end = 20
        self.mode = None
        self.events = None
        self.ch_name = ['No file']
        self.range_name = ['No file']
        self.range_name_buf = []
        self.ch_idx = 0
        self.range_idx = 0
        self.root = Tk()
        self.root.title("Расчет мощности")
        self.root.geometry("520x250")
        self.data = []
        self.data_standart = []
        self.file_name = ''
        self.root.resizable(width=False, height=False)
        self.btn_plot = Button(text="Рассчитать СПМ", command=self.calc, relief=RAISED)
        self.btn_plot.place(x=20, y=70)
        self.btn_plot.config(state="disabled")
        self.r_var = BooleanVar()

        self.r1 = Radiobutton(text='По умолчанию', variable=self.r_var, value=0, command=self.select_mode)
        self.r2 = Radiobutton(text='Расширенные\nнастройки', variable=self.r_var, value=1, command=self.select_mode)
        self.r1.place(x=20, y=140)
        self.r2.place(x=20, y=160)

        self.listbox = Listbox(self.root, selectmode=MULTIPLE, exportselection=0, selectbackground='BLUE')
        self.listbox.place(x=240, y=30)
        self.listbox.config(width=10)
        self.listbox_range = Listbox(self.root, selectmode=MULTIPLE, exportselection=0, selectbackground='BLUE')
        self.listbox_range.place(x=340, y=30)
        self.listbox_range.config(width=10)

        self.entry = Entry(self.root)
        self.entry.place(x=430, y=30)
        self.entry.config(width=9)
        self.btn_open = Button(text="Открыть файл", command=lambda: [f() for f in [self.load_file, self.load_plx]],
                               relief=RAISED)
        self.btn_open.place(x=20, y=30)
        self.btn_open.config(width=13)
        self.status = Label(text="Файл не выбран")
        self.status.place(x=20, y=200)
        self.btn_add = Button(text="Добавить", command=self.addItem,
                              relief=RAISED)
        self.btn_add.place(x=430, y=60)
        self.btn_add.config(width=6)
        self.btn_add.config(height=1)
        self.btn_del = Button(text="Удалить", command=self.delList,
                              relief=RAISED)
        self.btn_del.place(x=430, y=90)
        self.btn_del.config(width=6)
        self.btn_del.config(height=1)

        self.i_var = IntVar()

        self.s1 = Radiobutton(text='1', variable=self.i_var, value=1, command=self.select_smell)
        self.s2 = Radiobutton(text='2', variable=self.i_var, value=2, command=self.select_smell)
        self.s3 = Radiobutton(text='3', variable=self.i_var, value=3, command=self.select_smell)
        self.s4 = Radiobutton(text='4', variable=self.i_var, value=4, command=self.select_smell)
        self.s5 = Radiobutton(text='5', variable=self.i_var, value=5, command=self.select_smell)
        self.s6 = Radiobutton(text='6', variable=self.i_var, value=6, command=self.select_smell)
        self.s1.place(x=430, y=120)
        self.s2.place(x=430, y=140)
        self.s3.place(x=430, y=160)
        self.s4.place(x=430, y=180)
        self.s5.place(x=430, y=200)
        self.s6.place(x=430, y=220)
        self.s1.select()
        self.v1 = StringVar(self.root, value='1')
        self.v2 = StringVar(self.root, value='20')
        self.start_w = Entry(self.root, textvariable=self.v1)
        self.start_w.place(x=180, y=30)
        self.start_w.config(width=3)
        self.end_w = Entry(self.root, textvariable=self.v2)
        self.end_w.place(x=180, y=50)
        self.end_w.config(width=3)
        #
        self.sec1 = StringVar(self.root, value='-5')
        self.sec2 = StringVar(self.root, value='5')
        self.start_sec = Entry(self.root, textvariable=self.sec1)
        self.start_sec.place(x=180, y=90)
        self.start_sec.config(width=3)
        self.end_sec = Entry(self.root, textvariable=self.sec2)
        self.end_sec.place(x=180, y=110)
        self.end_sec.config(width=3)
        self.textbox_name = Label(text="От")
        self.textbox_name.place(x=160, y=90)
        self.textbox_name = Label(text="До")
        self.textbox_name.place(x=160, y=110)
        self.textbox_name = Label(text="Секунды")
        self.textbox_name.place(x=160, y=70)
        self.av_sec = IntVar()
        self.av_stim = IntVar()
        self.av = Label(text="Усреднение:")
        self.av.place(x=147, y=141)
        self.av_sec_cb = Checkbutton(self.root, text='секунд', variable=self.av_sec)
        self.av_sec_cb.place(x=147, y=160)
        self.av_stim_cb = Checkbutton(self.root, text='стимулов', variable=self.av_stim)
        self.av_stim_cb.place(x=147, y=180)
        #
        self.textbox_name = Label(text="От")
        self.textbox_name.place(x=160, y=30)
        self.textbox_name = Label(text="До")
        self.textbox_name.place(x=160, y=50)

        self.btn_reset = Button(text="Сброс", command=self.reset,
                                relief=RAISED)
        self.btn_reset.place(x=20, y=110)
        self.btn_reset.config(state="disabled")
        self.btn_reset.config(width=13)
        self.load = Label(text="")
        self.load.place(x=20, y=220)
        self.textbox_name = Label(text="Каналы")
        self.textbox_name.place(x=240, y=10)
        self.textbox_name = Label(text="Диапазоны")
        self.textbox_name.place(x=340, y=10)
        self.textbox_name = Label(text="Стимулы")
        self.textbox_name.place(x=160, y=10)

        self.select_mode(self.r_var.get())

        self.root.mainloop()

    def select_smell(self):
        self.smell = self.i_var.get()
        return

    def select_mode(self, event=None):
        self.mode = self.r_var.get()
        freq_band = [[1, 4], [4, 8], [8, 14], [14, 30], [30, 60], [60, 90], [90, 170]]
        if self.mode == True:
            self.btn_add.config(state="normal")
            self.btn_del.config(state="normal")
            self.range_name = [str(i[0]) + '-' + str(i[1]) for i in freq_band]
            self.range_name_buf = freq_band
            self.av_sec_cb.config(state='normal')
            self.av_stim_cb.config(state='normal')
            self.start_sec.config(state='normal')
            self.end_sec.config(state='normal')
            self.start_w.config(state='normal')
            self.end_w.config(state='normal')
            self.g_i_range()
        else:
            self.listbox_range.delete(0, 'end')
            self.range_name = [str(i[0]) + '-' + str(i[1]) for i in freq_band]
            self.btn_add.config(state="disabled")
            self.btn_del.config(state="disabled")
            self.av_sec_cb.config(state='disabled')
            self.av_stim_cb.config(state='disabled')
            self.start_sec.config(state='disabled')
            self.end_sec.config(state='disabled')
            self.start_w.config(state='disabled')
            self.end_w.config(state='disabled')

    def addItem(self):
        try:
            result = re.search(r'[0-9]-[0-9]', self.entry.get())
            result.group(0)
            self.range_name_buf.append([int(self.entry.get().split('-')[0]), int(self.entry.get().split('-')[1])])
            self.range_name_buf = sorted(self.range_name_buf, key=lambda x: x[0])
            range_name_str = [str(i[0]) + '-' + str(i[1]) for i in self.range_name_buf]
            self.listbox_range.delete(0, 'end')
            for item in range_name_str:
                self.listbox_range.insert(END, item)
            self.entry.delete(0, END)
            self.listbox_range.select_set(0)  # This only sets focus on the first item.
            self.listbox_range.event_generate("<<ListboxSelect>>")
        except AttributeError as error:
            mb.showinfo("Некорректный ввод", "Пример ввода диапазона: 70-120")

    def delList(self):
        select = list(self.listbox_range.curselection())
        select.reverse()
        remove_list=[]
        for i in select:
            self.listbox_range.delete(i)
            remove_list.append(self.range_name_buf[i])
        self.range_name = []
        self.range_name_buf = [e for e in self.range_name_buf if e not in remove_list]
        self.listbox_range.select_set(0)  # This only sets focus on the first item.
        self.listbox_range.event_generate("<<ListboxSelect>>")

    def load_file(self):
        f_n = filedialog.askopenfilename(initialdir='./', title="Выбрать файл", filetypes=[("Plexon files", "*.plx")])

        if len(f_n) == 0:
            return
        self.btn_open.config(state="disabled")
        self.file_name = f_n
        self.btn_add.config(state="disabled")
        self.btn_del.config(state="disabled")
        self.status.config(text=self.file_name)
        self.load.config(text="Загрузка")

        self.root.update()

    def load_plx(self):
        if self.file_name == '':
            return
        DECIMATE = 10
        segm = PlexonIO(self.file_name).read_segment()  # bool(take_spikes)load_spike_waveform=False
        signal = [asig for asig in segm.analogsignals if len(
            asig) > 0]  # каждый асиг имеет атрибут name - названия каналов, split по пробелу и take 1 element
        ch_names = [asig.annotations['channel_name'] for asig in segm.analogsignals if len(
            asig) > 0]
        signal = np.asarray(np.multiply(np.concatenate(signal, axis=1), 1000),
                            np.float32)  # переводим из миливольт в микровольты
        eeg = np.asarray(signal[::DECIMATE])
        eeg_standart = np.asarray(signal[::DECIMATE])
        self.data_standart = eeg_standart
        # if self.mode==True:
        LEN_STIMUL_SECS, BEFORE_STIMUL, LEN_STIMUL = 5, 5000, 5000
        all_valves_events = [ev for ev in segm.events if len(ev) > 7]
        if len(all_valves_events) > 0:
            # отбираем массивы равные первому
            len_first_arr = len(all_valves_events[0])
            valves_events = [ev for ev in all_valves_events if len(ev) == len_first_arr]
            if len(valves_events) == 1:
                valves_events = all_valves_events[1:]
            else:
                raise Exception('Import .plx file error!')
        else:
            valves_events = []
        sampling_rate = int(segm.analogsignals[-1].sampling_rate)
        num_cutoffs = signal.shape[0]
        stimul = np.zeros(num_cutoffs, dtype=np.int16)
        for i_valve, events in enumerate(valves_events):  # по клапанам
            for event in events:  # по событиям
                begin = int(event * sampling_rate)
                stimul[begin: begin + LEN_STIMUL_SECS * sampling_rate] = 1 + i_valve
        stimul = stimul[::DECIMATE]
        stimul_before = np.delete(stimul, np.s_[:BEFORE_STIMUL])
        stimul_before = np.append(stimul_before, np.zeros((BEFORE_STIMUL)))
        INCLUDE_STIMUL_VALUES = [i_valve + 1 for i_valve, events in enumerate(valves_events)]
        if len(INCLUDE_STIMUL_VALUES) == 6:
            mask = np.any([[(stimul == INCLUDE_STIMUL_VALUES[0])], [(stimul == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul == INCLUDE_STIMUL_VALUES[2])], [(stimul == INCLUDE_STIMUL_VALUES[3])],
                           [(stimul == INCLUDE_STIMUL_VALUES[4])], [(stimul == INCLUDE_STIMUL_VALUES[5])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[0])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[2])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[3])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[4])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[5])]], axis=0)[
                0]  # [0] because of 2d result array
        elif len(INCLUDE_STIMUL_VALUES) == 5:
            mask = np.any([[(stimul == INCLUDE_STIMUL_VALUES[0])], [(stimul == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul == INCLUDE_STIMUL_VALUES[2])], [(stimul == INCLUDE_STIMUL_VALUES[3])],
                           [(stimul == INCLUDE_STIMUL_VALUES[4])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[0])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[2])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[3])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[4])]], axis=0)[
                0]  # [0] because of 2d result array
        elif len(INCLUDE_STIMUL_VALUES) == 4:
            mask = np.any([[(stimul == INCLUDE_STIMUL_VALUES[0])], [(stimul == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul == INCLUDE_STIMUL_VALUES[2])], [(stimul == INCLUDE_STIMUL_VALUES[3])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[0])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[2])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[3])]], axis=0)[
                0]  # [0] because of 2d result array
        elif len(INCLUDE_STIMUL_VALUES) == 3:
            mask = np.any([[(stimul == INCLUDE_STIMUL_VALUES[0])], [(stimul == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul == INCLUDE_STIMUL_VALUES[2])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[0])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[2])]], axis=0)[
                0]  # [0] because of 2d result array
        elif len(INCLUDE_STIMUL_VALUES) == 2:
            mask = np.any([[(stimul == INCLUDE_STIMUL_VALUES[0])], [(stimul == INCLUDE_STIMUL_VALUES[1])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[0])],
                           [(stimul_before == INCLUDE_STIMUL_VALUES[1])]], axis=0)[
                0]  # [0] because of 2d result array
        elif len(INCLUDE_STIMUL_VALUES) == 1:
            mask = stimul == INCLUDE_STIMUL_VALUES[0]
        sample_size = LEN_STIMUL + BEFORE_STIMUL

        eeg = eeg[mask]
        stimul = stimul[mask]
        a = float(len(stimul)) % sample_size
        n = len(stimul) // sample_size
        if a != 0:  # обрежем массив
            b = n * sample_size
            eeg = eeg[:b, :]
            stimul = stimul[:b]
        n = len(eeg) // sample_size
        eeg = (np.reshape(eeg, (n, sample_size, eeg.shape[1]))).astype(np.float32).transpose((0, 2, 1))

        stimul = np.reshape(stimul, (n, sample_size))[:, -1].flatten()
        # eeg = eeg
        rad_but_smell = [self.s1, self.s2, self.s3, self.s4, self.s5, self.s6]

        dif = set(range(1,7)) - set(np.unique(stimul))

        for i in dif:
            rad_but_smell[i-1].config(state="disabled")

        self.ch_name = ch_names  # np.where((np.in1d(ch_names, self.ch_name)*1)==1)[0]
        # self.range_name =
        self.g_i()
        self.data = eeg
        self.events = stimul
        self.load.config(text="")
        self.btn_plot.config(state="normal")
        self.btn_add.config(state="normal")
        self.btn_del.config(state="normal")
        return

    def select_item(self, *args):
        value_name = [self.listbox.get(idx) for idx in self.listbox.curselection()]
        value_idx = [idx for idx in self.listbox.curselection()]
        if len(value_name) == 0:
            self.listbox.select_set(0)
            value_name = [self.listbox.get(0)]
            value_idx = [0]
        self.ch_name = value_name
        self.ch_idx = value_idx
        self.btn_reset.config(state="normal")
        return value_name, value_idx

    def select_item_range(self, *args):
        value_name_range = [self.listbox_range.get(idx) for idx in self.listbox_range.curselection()]
        value_idx_range = [idx for idx in self.listbox_range.curselection()]
        if len(value_name_range) == 0:
            self.listbox_range.select_set(0)
            value_name_range = [self.listbox_range.get(0)]
        self.range_name = value_name_range
        self.range_idx = value_idx_range
        return value_name_range, value_idx_range

    def g_i(self):
        for item in self.ch_name:
            self.listbox.insert(END, item)
        self.listbox.bind('<<ListboxSelect>>', self.select_item)
        self.listbox.select_set(0)
        self.listbox.event_generate("<<ListboxSelect>>")

    def g_i_range(self):
        for item in self.range_name:
            self.listbox_range.insert(END, item)
        self.listbox_range.bind('<<ListboxSelect>>', self.select_item_range)
        self.listbox_range.select_set(0)
        self.listbox_range.event_generate("<<ListboxSelect>>")

    def reset(self):
        self.listbox.delete(0, 'end')
        self.btn_open.config(state="normal")
        self.btn_plot.config(state="disabled")
        self.status.config(text='Файл не выбран')
        self.file_name = ''
        self.btn_reset.config(state="disabled")
        for i in range(1,7):
            rad_but_smell[i-1].config(state="normal")
        return

    def calc(self):
        self.start = int(self.start_w.get())
        self.end = int(self.end_w.get())
        if isinstance(self.range_name[0][0], str):
            self.range_name = [[int(i.split('-')[0]), int(i.split('-')[1])] for i in self.range_name if i[0]]

        ch_names = self.ch_name
        ch_idx = self.ch_idx
        # self.listbox.bindtags((self.listbox, "all"))

        STEP = 5000
        SAMPLE_RATE = 1000

        if not self.mode:
            data = self.data_standart
            data = data[:, ch_idx]
            freq_bank = ((1, 4), (4, 8), (8, 14), (14, 30), (30, 60), (60, 90), (90, 170))
            df, pxxs = od(), []
            data = data[:-1 * (data.shape[0] % STEP)]
            for i in range(data.shape[0] // STEP):
                f, pxx = welch(data[i * STEP: (i + 1) * STEP], fs=SAMPLE_RATE, axis=0, nperseg=1024)
                pxxs.append(pxx)
            for band in freq_bank:
                names = [ch_name + ' ' + str(band[0]) + '-' + str(band[1]) + ' Hz' for ch_name in ch_names]
                for ch, name in enumerate(names):
                    df[name] = [np.sum(pxx[(f >= band[0]) & (f <= band[1])][:, ch]) for pxx in pxxs]
            df = pd.DataFrame(df)
            savefile = tkFileDialog.asksaveasfilename(title="Сохранить файл",
                                                      filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
            if savefile != '':
                if not resname.endswith(".xlsx"):
                    resname+='.xlsx'
                fname = savefile
                workbook = xlsxwriter.Workbook(fname)
                workbook.close()
                writer = pd.ExcelWriter(fname)
                df.to_excel(writer, index=False)
                writer.save()
                self.listbox.delete(0, 'end')
                self.btn_open.config(state="normal")
                self.btn_plot.config(state="disabled")
                self.status.config(text='Файл не выбран')
        else:
            start_sec, end_sec = int(self.start_sec.get()), int(self.end_sec.get())
            true_stim_len = self.data.shape[2] // 2 // SAMPLE_RATE
            start_time_idx, end_time_idx = (true_stim_len + start_sec) * SAMPLE_RATE, (true_stim_len + end_sec) * SAMPLE_RATE
            if start_time_idx < 0:
                start_time_idx = 0
            if end_time_idx > self.data.shape[2]:
                end_time_idx = self.data.shape[2]
            print(self.data.shape)
            data = self.data[self.events == self.smell][self.start - 1: self.end, ch_idx, start_time_idx: end_time_idx]
            print(data.shape)
            freq_bank = sorted(self.range_name)
            av_sec = bool(self.av_sec.get())
            av_stim = bool(self.av_stim.get())
            ts = od()
            exps = [0] if av_stim else range(len(data))
            if av_stim:
                zero_cell = ''
            else:
                zero_cell = '{} – {} c'.format(start_sec, end_sec)
            if av_sec:
                it = [[0, end_time_idx]]
                if av_stim:
                    secs = np.array('{} – {} c'.format(start_sec, end_sec)).reshape((1, 1))
                else:
                    secs = np.array(list(map(str, range(1, len(data) + 1)))).reshape((1, -1))
            else:
                secs = np.array(['{} c'.format(t) for t in np.arange(start_sec, end_sec).astype(int)] * len(exps)).reshape((1, -1))
                it = [[i * SAMPLE_RATE, (i + 1) * SAMPLE_RATE] for i in range(data.shape[2] // SAMPLE_RATE)]
                if not av_stim:
                    stim_nums = np.array([[' ({})'.format(i)] * len(it) for i in range(1, len(exps) + 1)]).reshape((1, -1))
                    secs = np.array([secs[0, i] + stim_nums[0, i] for i in range(secs.size)])
            for exp in exps:
                for n in it:
                    if av_stim:
                        f, pxx = welch(data[:, :, n[0]: n[1]], fs=SAMPLE_RATE, axis=2, nperseg=1000)
                        pxx = np.mean(pxx, 0)
                    else:
                        f, pxx = welch(data[exp, :, n[0]: n[1]], fs=SAMPLE_RATE, axis=1, nperseg=1000)
                    for ch, ch_name in enumerate(ch_names):
                        fr_ch = []
                        for rhythm in freq_bank[::-1]:
                            fr_ch.append(np.sum(pxx[ch, (f <= rhythm[1]) & (f >= rhythm[0])]))
                        ts[ch_name] = np.hstack((ts.get(ch_name, np.empty((len(freq_bank), 0))), np.array(fr_ch).reshape((-1, 1))))

            freq_names = [str(band[0]) + '-' + str(band[1]) + ' Гц' for band in freq_bank[::-1]]
            for ch in ts:
                col1 = np.array([zero_cell] + freq_names).reshape((-1, 1))
                ts[ch] = np.hstack((col1, np.vstack((secs, ts[ch]))))
                ks, vs = ts[ch].T[:, 0], ts[ch].T[:, 1:]
                ts[ch] = od(zip(ks, vs))

            resname = filedialog.asksaveasfilename(title="Сохранить файл",
                                                     filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
            if resname != '':
                if not resname.endswith(".xlsx"):
                    resname+='.xlsx'
                workbook = xlsxwriter.Workbook(resname)
                workbook.close()
                workbook1 = openpyxl.load_workbook(resname)
                std = workbook1.get_sheet_by_name('Sheet1')
                workbook1.remove_sheet(std)
                writer = pd.ExcelWriter(resname, engine='openpyxl')
                writer.book = workbook1
                for ch in ts:
                    df = pd.DataFrame(ts[ch])
                    df.to_excel(writer, sheet_name=ch, index=False)
                writer.save()
                writer.close()


A = app()
